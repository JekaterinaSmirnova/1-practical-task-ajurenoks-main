from openpyxl import Workbook, load_workbook 
wb=load_workbook('tests/test1.xlsx')
ws=wb.active
total=0
#write your code here
for rinda in range(2,ws.max_row+1):
    stundas=ws['C'+str(rinda)].value
    likme=ws['B'+str(rinda)].value
    if isinstance(stundas,(int,float)) and isinstance(likme,(int,float)):
         alga=stundas*likme
         if alga>3000:
             total=total+1
print(total)
wb.close()
